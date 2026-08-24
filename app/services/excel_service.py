from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from app.config import settings
from app.database.models.result import Result
from app.database.models.test import Question


class ExcelService:
    @staticmethod
    def parse_questions_from_excel(file_path: Path) -> Tuple[List[Dict[str, Any]], List[str]]:
        """
        Parses questions from an Excel file (.xlsx).
        Expected columns in first sheet:
        question | option_a | option_b | option_c | option_d | correct_answer | points

        Returns: (parsed_questions, errors_list)
        """
        errors: List[str] = []
        questions: List[Dict[str, Any]] = []

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            sheet = wb.active
            if sheet is None:
                return [], ["Excel faylida faol sahifa topilmadi."]

            # Find header mapping or assume row 1 is header
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                return [], ["Excel fayli bo'sh."]

            header = [str(col).strip().lower() if col is not None else "" for col in rows[0]]

            # Map column indexes
            col_map = {}
            aliases = {
                "question": ["question", "savol", "savol matni", "savollar", "question_text"],
                "option_a": ["option_a", "a", "a variant", "a)", "variant_a"],
                "option_b": ["option_b", "b", "b variant", "b)", "variant_b"],
                "option_c": ["option_c", "c", "c variant", "c)", "variant_c"],
                "option_d": ["option_d", "d", "d variant", "d)", "variant_d"],
                "correct_answer": ["correct_answer", "correct", "to'g'ri javob", "togri javob", "kalit", "javob", "answer"],
                "points": ["points", "ball", "baho", "ballar", "point"]
            }

            for key, names in aliases.items():
                for idx, col_name in enumerate(header):
                    if col_name in names:
                        col_map[key] = idx
                        break

            # If headers didn't match by name, fallback to positional if columns >= 6
            if len(col_map) < 6:
                col_map = {
                    "question": 0,
                    "option_a": 1,
                    "option_b": 2,
                    "option_c": 3,
                    "option_d": 4,
                    "correct_answer": 5,
                    "points": 6 if len(header) > 6 else -1
                }

            # Cyrillic lookalikes map for correct answers
            CYRILLIC_LOOKALIKES = {
                "А": "A", "В": "B", "С": "C", "D": "D",
                "а": "A", "в": "B", "с": "C", "d": "D",
                "1": "A", "2": "B", "3": "C", "4": "D"
            }

            # Parse data rows (start from row 2)
            for row_idx, row in enumerate(rows[1:], start=2):
                if not any(row):
                    continue  # skip completely empty row

                q_text = row[col_map["question"]] if col_map.get("question", -1) < len(row) and col_map.get("question", -1) >= 0 else None
                opt_a = row[col_map["option_a"]] if col_map.get("option_a", -1) < len(row) and col_map.get("option_a", -1) >= 0 else None
                opt_b = row[col_map["option_b"]] if col_map.get("option_b", -1) < len(row) and col_map.get("option_b", -1) >= 0 else None
                opt_c = row[col_map["option_c"]] if col_map.get("option_c", -1) < len(row) and col_map.get("option_c", -1) >= 0 else None
                opt_d = row[col_map["option_d"]] if col_map.get("option_d", -1) < len(row) and col_map.get("option_d", -1) >= 0 else None
                correct = row[col_map["correct_answer"]] if col_map.get("correct_answer", -1) < len(row) and col_map.get("correct_answer", -1) >= 0 else None
                points_val = row[col_map["points"]] if col_map.get("points", -1) < len(row) and col_map.get("points", -1) >= 0 else 1.0

                # Validate Question Text
                if not q_text or not str(q_text).strip():
                    continue

                # Validate Options
                if not opt_a or not opt_b or not opt_c or not opt_d:
                    errors.append(f"❌ {row_idx}-qator: Barcha 4 ta variant (A, B, C, D) to'liq bo'lishi kerak.")
                    continue

                # Validate Correct Answer (normalize Cyrillic / lowercase)
                raw_correct = str(correct).strip() if correct else ""
                correct_str = CYRILLIC_LOOKALIKES.get(raw_correct, raw_correct.upper())
                if correct_str not in ["A", "B", "C", "D"]:
                    errors.append(f"❌ {row_idx}-qator: To'g'ri javob noto'g'ri ('{raw_correct}'). Faqat A, B, C yoki D bo'lishi kerak.")
                    continue

                # Validate Points
                points = 1.0
                if points_val is not None:
                    try:
                        points = float(points_val)
                        if points <= 0:
                            points = 1.0
                    except (ValueError, TypeError):
                        points = 1.0

                questions.append({
                    "text": str(q_text).strip(),
                    "option_a": str(opt_a).strip(),
                    "option_b": str(opt_b).strip(),
                    "option_c": str(opt_c).strip(),
                    "option_d": str(opt_d).strip(),
                    "correct_option": correct_str,
                    "points": points
                })

            wb.close()
        except Exception as e:
            errors.append(f"Faylni o'qishda xatolik: {str(e)}")

        return questions, errors

    @staticmethod
    def generate_sample_questions_template() -> Path:
        """
        Generates a beautiful ready-to-use sample questions template (.xlsx).
        """
        settings.EXCEL_DIR.mkdir(parents=True, exist_ok=True)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Savollar"

        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
        alignment_center = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin', color='CBD5E0'),
            right=Side(style='thin', color='CBD5E0'),
            top=Side(style='thin', color='CBD5E0'),
            bottom=Side(style='thin', color='CBD5E0')
        )

        headers = [
            "Savol matni (question)",
            "A varianti (option_a)",
            "B varianti (option_b)",
            "C varianti (option_c)",
            "D varianti (option_d)",
            "To'g'ri javob (correct_answer)",
            "Ball (points)"
        ]
        ws.append(headers)

        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment_center

        sample_rows = [
            ["O'zbekiston poytaxti qaysi shahar?", "Samarqand", "Toshkent", "Buxoro", "Xiva", "B", 1.0],
            ["2 + 2 * 2 ifodaning qiymati nechaga teng?", "6", "8", "4", "10", "A", 1.0],
            ["Alisher Navoiy qaysi asr mutafakkiri?", "XIV asr", "XV asr", "XVI asr", "XIII asr", "B", 1.5],
            ["Suvning kimyoviy formulasi qaysi?", "CO2", "NaCl", "H2O", "O2", "C", 1.0],
            ["Eng katta okean qaysi?", "Atlantika", "Hind", "Shimoliy Muz", "Tinch okeani", "D", 1.0]
        ]

        for idx, row in enumerate(sample_rows, start=2):
            ws.append(row)
            for col_num in range(1, len(row) + 1):
                cell = ws.cell(row=idx, column=col_num)
                cell.border = thin_border
                if col_num in [6, 7]:
                    cell.alignment = alignment_center

        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 15)

        file_path = settings.EXCEL_DIR / "Savollar_Namuna_Shablon.xlsx"
        wb.save(file_path)
        wb.close()
        return file_path


    @staticmethod
    def export_results_to_excel(results: List[Result], test_title: str) -> Path:
        """
        Exports test results to a professional 2-sheet Excel spreadsheet:
        Sheet 1: Umumiy Natijalar (Rank, Name, Phone, School, Grade, Score, %, Times)
        Sheet 2: Savollar Matritsasi (Question-by-Question breakdown: 1:A+, 2:B-)
        """
        wb = openpyxl.Workbook()

        # --- SHEET 1: Umumiy Natijalar ---
        ws1 = wb.active
        ws1.title = "Umumiy Natijalar"

        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
        alignment_center = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin', color='CBD5E0'),
            right=Side(style='thin', color='CBD5E0'),
            top=Side(style='thin', color='CBD5E0'),
            bottom=Side(style='thin', color='CBD5E0')
        )

        headers1 = [
            "№",
            "Ism",
            "Familiya",
            "Telefon",
            "Maktab / Muassasa",
            "Sinf / Toifa",
            "Username",
            "Test nomi",
            "To'g'ri javob",
            "Xato javob",
            "To'plangan ball",
            "Ko'rsatkich (%)",
            "Boshlangan vaqti",
            "Yakunlangan vaqti"
        ]

        ws1.append(headers1)

        for col_num, _ in enumerate(headers1, 1):
            cell = ws1.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment_center

        max_questions = 0
        correct_keys_map = {}

        for idx, res in enumerate(results, start=1):
            user = res.user
            attempt = res.attempt
            started = attempt.started_at.strftime("%d.%m.%Y %H:%M") if attempt and attempt.started_at else ""
            finished = attempt.finished_at.strftime("%d.%m.%Y %H:%M") if attempt and attempt.finished_at else ""

            if attempt and attempt.option_order:
                corr_k = attempt.option_order.get("correct_keys", {})
                if len(corr_k) > max_questions:
                    max_questions = len(corr_k)
                    correct_keys_map = {int(k): v for k, v in corr_k.items()}

            row_data = [
                idx,
                user.first_name if user else "-",
                user.last_name if user else "-",
                user.phone_number if user and user.phone_number else "-",
                user.school if user and user.school else "-",
                user.grade if user and user.grade else "-",
                f"@{user.username}" if user and user.username else "-",
                test_title,
                res.correct_count,
                res.incorrect_count,
                res.total_score,
                f"{res.percentage}%",
                started,
                finished
            ]
            ws1.append(row_data)

            for col_num in range(1, len(row_data) + 1):
                cell = ws1.cell(row=idx + 1, column=col_num)
                cell.border = thin_border
                if col_num in [1, 9, 10, 11, 12, 13, 14]:
                    cell.alignment = alignment_center

        for col in ws1.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws1.column_dimensions[col_letter].width = max(max_len + 3, 12)

        # --- SHEET 2: Savollar Matritsasi (Tahlil) ---
        if max_questions > 0:
            ws2 = wb.create_sheet(title="Savollar Matritsasi")
            headers2 = ["№", "F.I.Sh", "Ball", "Foiz (%)"]
            for q_idx in range(1, max_questions + 1):
                corr_val = correct_keys_map.get(q_idx, "")
                suffix = f" ({corr_val})" if corr_val else ""
                headers2.append(f"{q_idx}-savol{suffix}")

            ws2.append(headers2)
            header_fill2 = PatternFill(start_color="2B6CB0", end_color="2B6CB0", fill_type="solid")

            for col_num, _ in enumerate(headers2, 1):
                cell = ws2.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill2
                cell.alignment = alignment_center

            from app.services.scoring_service import ScoringService

            for idx, res in enumerate(results, start=1):
                user = res.user
                attempt = res.attempt
                user_full = f"{user.first_name or ''} {user.last_name or ''}".strip() if user else "O'quvchi"
                
                row_m = [idx, user_full, res.total_score, f"{res.percentage}%"]

                user_answers = {}
                if attempt and attempt.option_order and "user_answers" in attempt.option_order:
                    raw_ua = attempt.option_order["user_answers"]
                    user_answers = {int(k): v for k, v in raw_ua.items()}

                for q_idx in range(1, max_questions + 1):
                    u_val = user_answers.get(q_idx)
                    corr_val = correct_keys_map.get(q_idx)
                    if u_val is not None:
                        is_corr = ScoringService.are_answers_equivalent(u_val, corr_val)
                        badge = f"{u_val} (+)" if is_corr else f"{u_val} (-)"
                    else:
                        badge = "—"
                    row_m.append(badge)

                ws2.append(row_m)

                for col_num in range(1, len(row_m) + 1):
                    cell = ws2.cell(row=idx + 1, column=col_num)
                    cell.border = thin_border
                    if col_num in [1, 3, 4] or col_num > 4:
                        cell.alignment = alignment_center

            for col in ws2.columns:
                max_len = max(len(str(cell.value or "")) for cell in col)
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                ws2.column_dimensions[col_letter].width = max(max_len + 3, 11)

        file_name = f"results_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        output_path = settings.EXCEL_DIR / file_name
        wb.save(output_path)
        wb.close()
        return output_path

    @staticmethod
    def export_users_to_excel(users: list) -> Path:
        settings.EXCEL_DIR.mkdir(parents=True, exist_ok=True)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Foydalanuvchilar"

        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
        alignment_center = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin', color='D0D0D0'),
            right=Side(style='thin', color='D0D0D0'),
            top=Side(style='thin', color='D0D0D0'),
            bottom=Side(style='thin', color='D0D0D0')
        )

        headers = [
            "№",
            "Ism",
            "Familiya",
            "Telegram Username",
            "Telegram ID",
            "Telefon raqam",
            "Maktab / Muassasa",
            "Sinf / Guruh",
            "Holati",
            "Ro'yxatdan o'tgan sana"
        ]
        ws.append(headers)

        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment_center

        for idx, u in enumerate(users, start=1):
            status_str = "Bloklangan" if u.is_blocked else "Faol"
            created_str = u.created_at.strftime("%d.%m.%Y %H:%M") if u.created_at else ""

            row_data = [
                idx,
                u.first_name or "",
                u.last_name or "",
                f"@{u.username}" if u.username else "",
                u.telegram_id,
                u.phone_number or "",
                u.school or "",
                u.grade or "",
                status_str,
                created_str
            ]
            ws.append(row_data)

            for col_num in range(1, len(row_data) + 1):
                ws.cell(row=idx + 1, column=col_num).border = thin_border

        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        file_name = f"users_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        output_path = settings.EXCEL_DIR / file_name
        wb.save(output_path)
        wb.close()
        return output_path

    @staticmethod
    def export_users_to_pdf(users: list) -> Path:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors

        settings.EXCEL_DIR.mkdir(parents=True, exist_ok=True)
        file_name = f"users_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        output_path = settings.EXCEL_DIR / file_name

        doc = SimpleDocTemplate(
            str(output_path),
            pagesize=landscape(A4),
            rightMargin=25,
            leftMargin=25,
            topMargin=25,
            bottomMargin=25
        )
        styles = getSampleStyleSheet()
        elements = []

        title_style = ParagraphStyle(
            'PDFUserTitle',
            parent=styles['Heading1'],
            fontName='Helvetica-Bold',
            fontSize=16,
            textColor=colors.HexColor("#0B1B3D"),
            alignment=1,
            spaceAfter=12
        )
        th_style = ParagraphStyle(
            'TH',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=8.5,
            textColor=colors.whitesmoke,
            alignment=1
        )
        td_style = ParagraphStyle(
            'TD',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=8,
            alignment=0
        )
        td_center = ParagraphStyle(
            'TDC',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=8,
            alignment=1
        )

        elements.append(Paragraph(f"FOYDALANUVCHILAR RO'YXATI (Jami: {len(users)} ta)", title_style))

        table_data = [[
            Paragraph("T/r", th_style),
            Paragraph("Ism-Familiya", th_style),
            Paragraph("Username", th_style),
            Paragraph("Telegram ID", th_style),
            Paragraph("Telefon", th_style),
            Paragraph("Maktab / Muassasa", th_style),
            Paragraph("Sinf", th_style),
            Paragraph("Holati", th_style),
            Paragraph("Sana", th_style)
        ]]

        for idx, u in enumerate(users, start=1):
            full_name = f"{u.first_name or ''} {u.last_name or ''}".strip() or "Noma'lum"
            uname = f"@{u.username}" if u.username else "-"
            phone = u.phone_number or "-"
            school = u.school or "-"
            grade = u.grade or "-"
            status = "Bloklangan" if u.is_blocked else "Faol"
            created = u.created_at.strftime("%d.%m.%Y") if u.created_at else "-"

            table_data.append([
                Paragraph(str(idx), td_center),
                Paragraph(full_name, td_style),
                Paragraph(uname, td_style),
                Paragraph(str(u.telegram_id), td_center),
                Paragraph(phone, td_center),
                Paragraph(school, td_style),
                Paragraph(grade, td_center),
                Paragraph(status, td_center),
                Paragraph(created, td_center)
            ])

        # Widths: A4 landscape = 841.89, margins 50, usable = 791.89
        col_widths = [30, 130, 95, 80, 85, 175, 55, 65, 65]
        t = Table(table_data, colWidths=col_widths)
        t_style = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#0B1B3D")),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('BOX', (0, 0), (-1, -1), 1, colors.HexColor("#CBD5E1")),
            ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0"))
        ]
        for i in range(1, len(table_data)):
            bg = colors.HexColor("#F8FAFC") if i % 2 == 0 else colors.white
            t_style.append(('BACKGROUND', (0, i), (-1, i), bg))

        t.setStyle(TableStyle(t_style))
        elements.append(t)

        doc.build(elements)
        return output_path

    @staticmethod
    def export_leaderboard_to_pdf(students: list[dict]) -> Path:
        from reportlab.lib.pagesizes import A4, portrait
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors

        settings.EXCEL_DIR.mkdir(parents=True, exist_ok=True)
        file_name = f"leaderboard_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        output_path = settings.EXCEL_DIR / file_name

        doc = SimpleDocTemplate(
            str(output_path),
            pagesize=portrait(A4),
            rightMargin=30,
            leftMargin=30,
            topMargin=30,
            bottomMargin=30
        )
        styles = getSampleStyleSheet()
        elements = []

        title_style = ParagraphStyle(
            'LdTitle',
            parent=styles['Heading1'],
            fontName='Helvetica-Bold',
            fontSize=15,
            textColor=colors.HexColor("#0B1B3D"),
            alignment=1,
            spaceAfter=12
        )
        th_style = ParagraphStyle(
            'TH',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=8.5,
            textColor=colors.whitesmoke,
            alignment=1
        )
        td_style = ParagraphStyle(
            'TD',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=8,
            alignment=0
        )
        td_center = ParagraphStyle(
            'TDC',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=8,
            alignment=1
        )

        elements.append(Paragraph(f"UMUMIY PLATFORMA REYTINGI (Jami: {len(students)} ta)", title_style))

        table_data = [[
            Paragraph("O'rin", th_style),
            Paragraph("Ism-Familiya", th_style),
            Paragraph("Maktab / Muassasa", th_style),
            Paragraph("Sinf", th_style),
            Paragraph("Testlar", th_style),
            Paragraph("O'rtacha", th_style),
            Paragraph("Jami Ball", th_style)
        ]]

        for idx, s in enumerate(students, start=1):
            full_name = f"{s.get('first_name') or ''} {s.get('last_name') or ''}".strip() or "Noma'lum"
            school = s.get('school') or "-"
            grade = s.get('grade') or "-"
            tests_cnt = str(s.get('tests_taken', 0))
            avg_pct = f"{s.get('avg_percentage', 0.0):.1f}%"
            total_sc = f"{s.get('total_score', 0.0):.1f}"

            table_data.append([
                Paragraph(str(idx), td_center),
                Paragraph(full_name, td_style),
                Paragraph(school, td_style),
                Paragraph(grade, td_center),
                Paragraph(tests_cnt, td_center),
                Paragraph(avg_pct, td_center),
                Paragraph(total_sc, td_center)
            ])

        # Usable width portrait A4 = 595 - 60 = 535 pt
        col_widths = [35, 130, 160, 45, 50, 55, 60]
        t = Table(table_data, colWidths=col_widths)
        t_style = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#0B1B3D")),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('BOX', (0, 0), (-1, -1), 1, colors.HexColor("#CBD5E1")),
            ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0"))
        ]
        for i in range(1, len(table_data)):
            bg = colors.HexColor("#F8FAFC") if i % 2 == 0 else colors.white
            t_style.append(('BACKGROUND', (0, i), (-1, i), bg))

        t.setStyle(TableStyle(t_style))
        elements.append(t)

        doc.build(elements)
        return output_path
