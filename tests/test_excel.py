from pathlib import Path
import openpyxl
import pytest
from app.config import settings
from app.database.models.result import Result, TestAttempt
from app.database.models.user import User
from app.services.excel_service import ExcelService


def test_excel_import_parsing(tmp_path: Path):
    excel_path = tmp_path / "test_import.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Savollar"

    # Header
    ws.append(["question", "option_a", "option_b", "option_c", "option_d", "correct_answer", "points"])

    # Valid row
    ws.append(["O'zbekiston poytaxti qaysi?", "Samarqand", "Toshkent", "Buxoro", "Xiva", "B", 2.0])
    # Invalid correct_answer row
    ws.append(["Quyosh qayerdan chiqadi?", "G'arb", "Sharq", "Shimol", "Janub", "Z", 1.0])
    # Empty question row
    ws.append(["", "A", "B", "C", "D", "A", 1.0])

    wb.save(excel_path)
    wb.close()

    questions, errors = ExcelService.parse_questions_from_excel(excel_path)

    assert len(questions) == 1
    assert questions[0]["text"] == "O'zbekiston poytaxti qaysi?"
    assert questions[0]["correct_option"] == "B"
    assert questions[0]["points"] == 2.0

    assert len(errors) == 2
    assert "correct_answer noto'g'ri" in errors[0]
    assert "Savol matni kiritilmagan" in errors[1]


def test_excel_export():
    u = User(
        id=1,
        first_name="Anvar",
        last_name="Saidov",
        username="anvar_s",
        telegram_id=12345,
        phone_number="+998901112233",
        school="Litsiy",
        grade="10"
    )
    res = Result(
        id=1,
        user_id=1,
        test_id=1,
        correct_count=18,
        incorrect_count=2,
        unanswered_count=0,
        total_score=36.0,
        max_score=40.0,
        percentage=90.0,
        time_spent_seconds=1200
    )
    res.user = u
    res.attempt = TestAttempt(
        id=1,
        test_id=1,
        user_id=1,
        started_at=u.created_at,
        finished_at=u.created_at
    )

    path = ExcelService.export_results_to_excel([res], "Ona tili va Adabiyot")
    assert path.exists()

    wb = openpyxl.load_workbook(path)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    # Kamida 2 qator bo'lishi kerak: header + 1 natija
    assert len(rows) >= 2, f"Kutilgan >= 2 qator, topildi: {len(rows)}"

    # Hozirgi kod 14 ustun chiqaradi:
    # 0:№  1:Ism  2:Familiya  3:Telefon  4:Maktab/Muassasa  5:Sinf/Toifa
    # 6:Username  7:Test nomi  8:To'g'ri javob  9:Xato javob
    # 10:To'plangan ball  11:Ko'rsatkich(%)  12:Boshlangan vaqti  13:Yakunlangan vaqti
    assert rows[1][1] == "Anvar",        f"Ism kutildi 'Anvar', lekin: {rows[1][1]}"
    assert rows[1][2] == "Saidov",       f"Familiya kutildi 'Saidov', lekin: {rows[1][2]}"
    # Foiz ustuni (11-indeks) "90.0%" yoki 90.0 bo'lishi mumkin
    pct_val = rows[1][11]
    assert pct_val is not None, "Foiz (11-ustun) None bo'lmasligi kerak"
    if isinstance(pct_val, str):
        assert "90" in pct_val, f"Foizda 90 bo'lishi kerak, topildi: {pct_val}"
    else:
        assert abs(float(pct_val) - 90.0) < 0.1, f"Foiz ~90 bo'lishi kerak, topildi: {pct_val}"
    wb.close()
